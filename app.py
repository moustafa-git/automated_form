from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook
from pathlib import Path
import os

app = Flask(__name__)

# === CONFIGURE THESE TO MATCH YOUR SETUP ===
# Use environment variable or default to relative path for deployment
BASE_DIR = Path(__file__).parent
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", BASE_DIR / "simple.xlsx"))
SHEET_NAME = os.getenv("SHEET_NAME", "SHOC")  # Data sheet name


def append_to_excel(form_data: dict):
    """
    Append one observation as a new row in the Excel tracker.
    Assumes the header row is already in the sheet.
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found at: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in workbook.")

    ws = wb[SHEET_NAME]

    # Find next empty row (based on existing data)
    # Sheet layout:
    #   Row 1: title
    #   Row 2: blank
    #   Row 3: header row (S/N, Date, Observer, ...)
    #   Row 4+: data rows
    next_row = ws.max_row + 1

    # Auto-generate serial number in column A (S/N)
    # If there is no data yet (only title + blank + header), start from 1.
    if next_row <= 4:
        serial_number = 1
    else:
        # Look at previous row S/N; if missing, fall back to row index minus header offset
        prev_sn = ws.cell(row=next_row - 1, column=1).value
        if isinstance(prev_sn, (int, float)):
            serial_number = int(prev_sn) + 1
        else:
            serial_number = max(1, next_row - 3)

    ws.cell(row=next_row, column=1, value=serial_number)

    # Mapping: match header row in SHOC sheet (row 3)
    # Column A: S/N (handled above)
    # Column B: Date
    # Column C: Observer
    # Column D: Company
    # Column E: Trade/Position
    # Column F: Observation Group
    # Column G: Location/Area
    # Column H: Observation Type
    # Column I: Observation
    # Column J: Action
    # Column K: Action Taken
    # Column L: Corrective Action / Preventive Action
    # Column M: Priority
    # Column N: Risk Rating
    # Column O: Custodian
    # Column P: Due Date
    # Column Q: Status
    # Column R: Comment

    ws.cell(row=next_row, column=2, value=form_data.get("date_occurred"))
    ws.cell(row=next_row, column=3, value=form_data.get("observer"))
    ws.cell(row=next_row, column=4, value=form_data.get("company"))
    ws.cell(row=next_row, column=5, value=form_data.get("trade_position"))
    ws.cell(row=next_row, column=6, value=form_data.get("observation_group"))
    ws.cell(row=next_row, column=7, value=form_data.get("location_area"))
    ws.cell(row=next_row, column=8, value=form_data.get("observation_type"))
    ws.cell(row=next_row, column=9, value=form_data.get("observation"))
    ws.cell(row=next_row, column=10, value=form_data.get("action"))
    ws.cell(row=next_row, column=11, value=form_data.get("action_taken"))
    ws.cell(row=next_row, column=12, value=form_data.get("corrective_preventive_action"))
    ws.cell(row=next_row, column=13, value=form_data.get("priority"))
    ws.cell(row=next_row, column=14, value=form_data.get("risk_rating"))
    ws.cell(row=next_row, column=15, value=form_data.get("custodian"))
    ws.cell(row=next_row, column=16, value=form_data.get("due_date"))
    ws.cell(row=next_row, column=17, value=form_data.get("status"))
    ws.cell(row=next_row, column=18, value=form_data.get("comment"))

    wb.save(EXCEL_PATH)


@app.route("/", methods=["GET"])
def index():
    return render_template("form.html")


@app.route("/submit", methods=["POST"])
def submit():
    form_data = {
        "title": request.form.get("title") or "",
        "safe_act": request.form.get("safe_act") or "",
        "positive_action": request.form.get("positive_action") or "",
        "unsafe_act": request.form.get("unsafe_act") or "",
        "immediate_corrective": request.form.get("immediate_corrective") or "",
        "preventive_action": request.form.get("preventive_action") or "",
        "observer": request.form.get("observer") or "",
        "company": request.form.get("company") or "",
        "date_occurred": request.form.get("date_occurred") or "",
        "location_area": request.form.get("location_area") or "",
        # extra tracker fields
        "trade_position": request.form.get("trade_position") or "",
        "observation_group": request.form.get("observation_group") or "",
        "observation_type": request.form.get("observation_type") or "",
        "observation": request.form.get("observation") or "",
        "action": request.form.get("action") or "",
        "action_taken": request.form.get("action_taken") or "",
        "corrective_preventive_action": request.form.get("corrective_preventive_action") or "",
        "priority": request.form.get("priority") or "",
        "risk_rating": request.form.get("risk_rating") or "",
        "custodian": request.form.get("custodian") or "",
        "due_date": request.form.get("due_date") or "",
        "status": request.form.get("status") or "",
        "comment": request.form.get("comment") or "",
    }

    append_to_excel(form_data)

    return redirect(url_for("index"))


if __name__ == "__main__":
    # For local testing. In production, use a proper WSGI server.
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=os.getenv("FLASK_DEBUG", "False").lower() == "true")


